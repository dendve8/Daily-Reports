from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from sqlalchemy import create_engine, Column, Integer, String, Text
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy import text
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Database setup
engine = create_engine('sqlite:///barang_keluar.db')
Base = declarative_base()

class BarangKeluar(Base):
    __tablename__ = 'barang_keluar'
    id = Column(Integer, primary_key=True)
    nama_barang = Column(Text)
    qty = Column(Integer)
    sesi = Column(String)
    lokasi = Column(String)
    time = Column(String)

if not os.path.exists("barang_keluar.db"):
    Base.metadata.create_all(engine)
else:
    with engine.connect() as connection:
        connection.execute(text("PRAGMA foreign_keys=ON;"))
        try:
            connection.execute(text("ALTER TABLE barang_keluar ADD COLUMN lokasi TEXT;"))
        except Exception:
            pass  # Kolom sudah ada

Session = sessionmaker(bind=engine)
db_session = Session()

# Fallback manual untuk nama hari
days_translation = {
    "Monday": "Senin",
    "Tuesday": "Selasa",
    "Wednesday": "Rabu",
    "Thursday": "Kamis",
    "Friday": "Jumat",
    "Saturday": "Sabtu",
    "Sunday": "Minggu"
}

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("📥 Input Data", callback_data='START_INPUT')],
        [InlineKeyboardButton("📤 Export Data Exe", callback_data='EXPORT')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "👋 Selamat datang di Bot Daily Report Abhinaya!\n\n"
        "Gunakan tombol di bawah ini untuk mulai input data atau ekspor laporan.",
        reply_markup=reply_markup
    )

async def input_barang(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("🌯 Breakfast", callback_data='BREAKFAST'),
            InlineKeyboardButton("🍵 Coffe Time 1", callback_data='Coffe Time 1'),
            InlineKeyboardButton("🍱 Lunch", callback_data='LUNCH'),
            InlineKeyboardButton("🍵 Coffe Time 2", callback_data='Coffe Time 2'),
        ],
        [
            InlineKeyboardButton("🍜 Dinner", callback_data='DINNER'),
            InlineKeyboardButton("🍵 Coffe Time 3", callback_data='Coffe Time 3'),
            InlineKeyboardButton("🍲 Supper", callback_data='SUPPER')
            InlineKeyboardButton("🍵 Coffe Time 4", callback_data='Coffe Time 4'),
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.user_data['input_data'] = {}
    await update.callback_query.edit_message_text("Pada Waktu :", reply_markup=reply_markup)

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data == 'START_INPUT':
        await input_barang(update, context)
    elif query.data == 'EXPORT':
        await export_excel(update.callback_query.message, context)
    elif 'sesi' not in context.user_data['input_data']:
        context.user_data['input_data']['sesi'] = query.data
        await query.edit_message_text("📋 Kamu ingin memasukkan barang apa?")
        return
    elif 'lokasi' not in context.user_data['input_data']:
        context.user_data['input_data']['lokasi'] = query.data
        now = datetime.now()
        date_time_str = now.strftime("%Y-%m-%d %H:%M:%S")
        day_name = days_translation[now.strftime("%A")]  # Terjemahan manual untuk nama hari
        
        input_data = context.user_data['input_data']
        input_data['time'] = f"{day_name}, {date_time_str}"  # Gabungkan nama hari dan waktu
        new_entry = BarangKeluar(
            nama_barang=input_data['nama_barang'],
            qty=input_data['qty'],
            sesi=input_data['sesi'],
            lokasi=input_data['lokasi'],
            time=input_data['time']
        )
        db_session.add(new_entry)
        db_session.commit()
        keyboard = [
            [InlineKeyboardButton("🔄 Input Lagi", callback_data='START_INPUT')],
            [InlineKeyboardButton("📤 Hasilkan Excel File", callback_data='EXPORT')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("✅ Succses! Data telah tersimpan.", reply_markup=reply_markup)

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if 'nama_barang' not in context.user_data['input_data']:
        context.user_data['input_data']['nama_barang'] = update.message.text
        await update.message.reply_text("Masukkan jumlah barang:")
    elif 'qty' not in context.user_data['input_data']:
        context.user_data['input_data']['qty'] = int(update.message.text)
        keyboard = [
            [InlineKeyboardButton("📍 ABG-J MINICAMP", callback_data='MINI-CAMP')],
            [InlineKeyboardButton("📍 ABG-B BASECAMP", callback_data='BASE-CAMP')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Di mana posisi barang akan diletakkan?", reply_markup=reply_markup)

async def export_excel(message, context: ContextTypes.DEFAULT_TYPE):
    data = db_session.query(BarangKeluar).all()
    if not data:
        await message.reply_text("❌ Tidak ada data untuk diekspor.")
        return

    rows = [{
        "No": entry.id,
        "Nama Barang": entry.nama_barang,
        "Qty": entry.qty,
        "Sesi": entry.sesi,
        "Lokasi": entry.lokasi,
        "Time": entry.time
    } for entry in data]
    df = pd.DataFrame(rows)

    file_path = "Abhinaya_Daily_Report.xlsx"
    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(file_path)

    await message.reply_document(document=open(file_path, "rb"), filename=file_path)

def main():
    application = Application.builder().token("7764690604:AAEBPpnH7qbw96KatjRuLC37JG8Sqz8qeYM").build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    application.add_handler(CallbackQueryHandler(handle_callback))

    application.run_polling()

if __name__ == '__main__':
    main()
